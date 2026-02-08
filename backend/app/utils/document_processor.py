"""
Document text extraction and chunking utilities
"""

import io
import mimetypes
from typing import List, Tuple
from pathlib import Path

# Document parsers
from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook

# LangChain for text chunking
from langchain.text_splitter import RecursiveCharacterTextSplitter
from app.core.config import CHUNK_SIZE, CHUNK_OVERLAP_PERCENT


class DocumentProcessor:
    """Extract text from documents and chunk it"""
    
    @staticmethod
    def extract_text(file_content: bytes, mime_type: str = None, filename: str = None) -> str:
        """
        Extract plain text from document file
        
        Args:
            file_content: File content as bytes
            mime_type: MIME type of the file
            filename: Filename (for extension detection)
            
        Returns:
            Extracted plain text
        """
        # Detect MIME type from filename if not provided
        if not mime_type and filename:
            mime_type, _ = mimetypes.guess_type(filename)
        
        # Determine file type from extension
        file_ext = None
        if filename:
            file_ext = Path(filename).suffix.lower()
        
        # Extract text based on file type
        if mime_type == "application/pdf" or file_ext == ".pdf":
            return DocumentProcessor._extract_from_pdf(file_content)
        elif mime_type in [
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/msword"
        ] or file_ext in [".docx", ".doc"]:
            return DocumentProcessor._extract_from_docx(file_content)
        elif mime_type == "text/plain" or file_ext == ".txt":
            return DocumentProcessor._extract_from_text(file_content)
        elif mime_type in [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel"
        ] or file_ext in [".xlsx", ".xls"]:
            return DocumentProcessor._extract_from_excel(file_content)
        else:
            # Try to read as text as fallback
            try:
                return DocumentProcessor._extract_from_text(file_content)
            except:
                raise ValueError(f"Unsupported file type: {mime_type or file_ext}")
    
    @staticmethod
    def _extract_from_pdf(file_content: bytes) -> str:
        """Extract text from PDF"""
        pdf_reader = PdfReader(io.BytesIO(file_content))
        text_parts = []
        for page in pdf_reader.pages:
            text = page.extract_text()
            if text.strip():
                text_parts.append(text)
        return "\n\n".join(text_parts)
    
    @staticmethod
    def _extract_from_docx(file_content: bytes) -> str:
        """Extract text from Word document"""
        doc = DocxDocument(io.BytesIO(file_content))
        text_parts = []
        for paragraph in doc.paragraphs:
            text = paragraph.text.strip()
            if text:
                text_parts.append(text)
        return "\n\n".join(text_parts)
    
    @staticmethod
    def _extract_from_text(file_content: bytes) -> str:
        """Extract text from plain text file"""
        try:
            return file_content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                return file_content.decode('latin-1')
            except UnicodeDecodeError:
                return file_content.decode('utf-8', errors='ignore')
    
    @staticmethod
    def _extract_from_excel(file_content: bytes) -> str:
        """Extract text from Excel file"""
        workbook = load_workbook(io.BytesIO(file_content), data_only=True)
        text_parts = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_text = []
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    sheet_text.append(row_text)
            if sheet_text:
                text_parts.append(f"Sheet: {sheet_name}\n" + "\n".join(sheet_text))
        return "\n\n".join(text_parts)
        
    
    @staticmethod
    def chunk_text(
        text: str,
        chunk_size: int = CHUNK_SIZE,
        overlap_percent: float = CHUNK_OVERLAP_PERCENT
    ) -> List[str]:
        """
        Split text into chunks with overlap using LangChain's RecursiveCharacterTextSplitter
        
        Args:
            text: Text to chunk
            chunk_size: Size of each chunk in characters (default: 500)
            overlap_percent: Overlap percentage (default: 0.2 = 20%)
            
        Returns:
            List of text chunks
        """
        if not text or not text.strip():
            return []
        
        # Calculate overlap in characters
        overlap = int(chunk_size * overlap_percent)  # 20% of 500 = 100 chars
        
        # Initialize RecursiveCharacterTextSplitter
        # This uses a hierarchical approach: tries to split by paragraphs, then sentences, 
        # then words, then characters - preserving semantic structure
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=overlap,
            length_function=len,
            separators=["\n\n", "\n", ". ", " ", ""]  # Try these separators in order
        )
        
        # Split the text into chunks
        chunks = text_splitter.split_text(text)
        
        # Filter out empty chunks
        return [chunk.strip() for chunk in chunks if chunk.strip()]

